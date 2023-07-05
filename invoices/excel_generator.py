import os
import shutil

from django.conf import settings
from openpyxl import load_workbook


def copy_style(src, dst):
    dst.font = src.font.copy()
    dst.alignment = src.alignment.copy()
    dst.border = src.border.copy()
    dst.fill = src.fill.copy()


def generate_xlsx(
    user,
    invoice,
    mats,
    k_data,
):
    user_folder = os.path.join(
        settings.MEDIA_ROOT, f"excel/usersangebots/{user.username}/"
    )
    template_folder = os.path.join(settings.MEDIA_ROOT, "excel/")

    # Create the directory if it doesn't exist
    if not os.path.exists(user_folder):
        os.makedirs(user_folder)
    filename = "Elektriker_example.xlsx"
    template_path = os.path.join(template_folder, f"{filename}")

    xlsx_path = os.path.join(user_folder, f"Angebot_{invoice.invoice_id}.xlsx")

    shutil.copy(template_path, xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    user = invoice.user
    user_email = f"{user.email}"
    invoiceid = f"{invoice.invoice_id}"

    head = f"{invoiceid}"
    ws.title = f"Angebot_{invoiceid}"  # type: ignore
    row = 1
    ws.cell(row=row, column=2, value=user_email)  # type: ignore
    row = 2
    ws.cell(row=row, column=2, value=head)  # type: ignore
    row = 3
    kunden_name = []
    for data in k_data:
        name = f"{data.kunden_name}"
        kunden_name.append(name)
    ws.cell(row=row, column=2, value=kunden_name[0])  # type: ignore

    # Zählerschrank
    row = 8
    for data in k_data:
        kunden_name = f"{data.kunden_name}"
        zahlerschrank_typ = f"{data.zahlerschranken}"

        ws.cell(row=row, column=2, value=zahlerschrank_typ)  # type: ignore
        row += 1
        # type: ignore
    # Kabel/Leitungen
    row = 19
    for position in mats:
        model_name = f"{position.position}"
        if (
            "Mantellleitung" in model_name
            or "Kabelschellen" in model_name
            or "Kabelkanal" in model_name
            or "H07-VK" in model_name
        ):
            kabel = f"{position.position}"
            kabel_length = str(position.quantity).strip("Decimal()")
            ws.cell(row=row, column=2, value=kabel)  # type: ignore
            ws.cell(row=row, column=3, value=kabel_length)  # type: ignore
            row += 1

        # Bestückungsmaterial Zählerschrank
    row = 55
    for position in mats:
        model_name = f"{position.position}"
        if (
            "SLS" in model_name
            or "Überspannungsschutz" in model_name
            or "Leitungsschutzschalter" in model_name
            or "Hauptleitungsabzweigklemmen" in model_name
        ):
            detail = f"{position.position}"
            detail_quantity = str(position.quantity).strip("Decimal()")
            ws.cell(row=row, column=2, value=detail)  # type: ignore
            ws.cell(row=row, column=3, value=detail_quantity)  # type: ignore
            row += 1

        # Kleinteile
    row = 83
    for position in mats:
        model_name = f"{position.position}"
        if (
            not "SLS" in model_name
            or not "Überspannungsschutz" in model_name
            or not "Leitungsschutzschalter" in model_name
            or not "Hauptleitungsabzweigklemmen" in model_name
            or not "Mantellleitung" in model_name
            or not "Kabelschellen" in model_name
            or not "Kabelkanal" in model_name
            or not "H07-VK" in model_name
        ):
            detail = f"{position.position}"
            detail_quantity = str(position.quantity).strip("Decimal()")
            ws.cell(row=row, column=2, value=detail)  # type: ignore
            ws.cell(row=row, column=3, value=detail_quantity)  # type: ignore
            row += 1

    xlsx_path = os.path.join(user_folder, f"Angebot_{invoice.invoice_id}.xlsx")
    wb.save(xlsx_path)

    return xlsx_path
